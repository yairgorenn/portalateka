import streamlit as st
import pandas as pd
import io
import os
from openpyxl.styles import Font, PatternFill, Alignment

# הגדרת אורך המק"ט הרצוי בפוריטי (מעודכן ל-9 ספרות)
SKU_LENGTH = 9

# הגדרות עמוד ועיצוב בסיסי לעברית עם האייקון (גלגל שיניים)
st.set_page_config(page_title="חברת אטקה - מתקן קבצי הזמנה", page_icon="⚙️", layout="centered")

# CSS מורחב ליישור מלא לימין של כל סוגי הטקסטים והכותרות בדף
st.markdown("""
    <style>
    .stApp {
        direction: RTL;
        text-align: right;
    }
    .stApp h1, .stApp h2, .stApp h3, .stApp p, .stApp label, .stApp span {
        text-align: right !important;
        direction: RTL !important;
    }
    [data-testid="stFileUploadDropzone"] {
        direction: RTL;
        text-align: right !important;
    }
    [data-testid="stFileUploadDropzone"] * {
        text-align: right !important;
        direction: RTL !important;
    }
    div.stButton > button:first-child {
        background-color: #2e7d32;
        color: white;
        width: 100%;
        font-size: 20px;
        font-weight: bold;
        padding: 12px;
        border-radius: 8px;
    }
    </style>
    """, unsafe_allow_html=True)

# --- כותרות וטקסט ---
st.title("⚙️ מערכת תיקון קבצי הזמנות")
st.write(
    "העלו את קובץ ההזמנה שלכם כאן. המערכת תסדר את האפסים במק\"ט ותבדוק תקינות של כמויות ומק\"טים.")

# רכיב העלאת קובץ
uploaded_file = st.file_uploader("בחרו קובץ אקסל (xlsx) או CSV", type=["xlsx", "csv"], label_visibility="collapsed")

if uploaded_file is not None:
    try:
        # קריאת הקובץ תוך התעלמות מוחלטת מהשורה הראשונה (skiprows=1)
        if uploaded_file.name.endswith('.csv'):
            df = pd.read_csv(uploaded_file, skiprows=1, header=None)
        else:
            df = pd.read_excel(uploaded_file, skiprows=1, header=None)

        # בדיקה שיש לפחות שתי עמודות בנתונים
        if df.shape[1] < 2:
            st.error("❌ שגיאה: הקובץ חייב להכיל לפחות שתי עמודות (מק\"ט וכמות).")
        else:
            # מערכים לשמירת הנתונים המעובדים ורשימות להתראות
            cleaned_skus = []
            cleaned_qtys = []
            qty_warnings = []
            sku_warnings = []

            # לולאה שעוברת שורה-שורה על הקובץ ומבצעת ולידציה מורחבת
            for idx, row in df.iterrows():
                # חישוב מספר השורה האמיתי באקסל (אינדקס 0 הוא שורה 2 באקסל המקורי בגלל skiprows=1)
                excel_row_num = idx + 2

                orig_sku = row[0]
                orig_qty = row[1]

                # --- א) בדיקת תקינות מק"ט ---
                if pd.isna(orig_sku) or str(orig_sku).strip() == "":
                    sku_val = ""
                    sku_warnings.append(f"שים לב מקט לא מלא בשורה {excel_row_num}")
                else:
                    sku_str = str(orig_sku).strip()
                    # ניקוי סיומות .0 שנוצרות לעיתים באקסל
                    if sku_str.endswith('.0'):
                        sku_str = sku_str[:-2]

                    # חוק חדש: פחות מ-7 תווים נשאר כמו שהוא ומוציא התראה
                    if len(sku_str) < 7:
                        sku_val = sku_str
                        sku_warnings.append(f"שים לב מקט לא מלא בשורה {excel_row_num}")
                    else:
                        # 7 תווים ומעלה - משלים עם אפסים משמאל ל-9 ספרות
                        sku_val = sku_str.zfill(SKU_LENGTH)

                # --- ב) בדיקת תקינות כמות ---
                qty_is_valid = False
                qty_val = orig_qty

                if not pd.isna(orig_qty):
                    try:
                        # בדיקה האם מדובר במספר (שלם או עשרוני)
                        qty_float = float(orig_qty)
                        # בדיקה האם המספר הוא שלם וגדול מ-0
                        if qty_float.is_integer() and int(qty_float) > 0:
                            qty_val = int(qty_float)
                            qty_is_valid = True
                    except ValueError:
                        pass  # השארת הערך המקורי כטקסט במידה ונכשל

                # אם הכמות לא תקינה (0, ריק, טקסט, או שבר) - משאירים את השורה ורושמים הערה
                if not qty_is_valid:
                    qty_warnings.append(f"שים לב בשורה מספר {excel_row_num} חסר כמות")
                    if pd.isna(orig_qty):
                        qty_val = ""

                cleaned_skus.append(sku_val)
                cleaned_qtys.append(qty_val)

            # בניית ה-DataFrame החדש עם השורות המלאות
            df_clean = pd.DataFrame({'מק"ט': cleaned_skus, 'כמות': cleaned_qtys})

            # --- הצגת ההתראות למשתמש במידה וקיימות ---
            all_warnings = qty_warnings + sku_warnings
            if all_warnings:
                with st.expander(f"⚠️ נמצאו {len(all_warnings)} הערות בקובץ (לחצו לצפייה)", expanded=True):
                    for warning in all_warnings:
                        st.warning(warning)

            # יצירת קובץ אקסל חדש בזיכרון
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df_clean.to_excel(writer, index=False)

                # הגדרות עיצוב אקסל (RTL, פונטים, צבעים)
                workbook = writer.book
                worksheet = workbook.active

                # הגדרות גיליון
                worksheet.views.sheetView[0].showGridLines = True
                worksheet.sheet_properties.pageSetUpPr.fitToPage = True
                worksheet.sheet_view.rightToLeft = True

                # הגדרת סגנונות (פונט Tahoma 14)
                font_tahoma_regular = Font(name='Tahoma', size=14)
                font_tahoma_header = Font(name='Tahoma', size=14, bold=True, color='FFFFFF')  # לבן
                fill_header = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')  # כחול כהה
                center_align = Alignment(horizontal='center', vertical='center')

                # החלת העיצוב על כל התאים בקובץ
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = center_align
                        if cell.row == 1:  # שורת הכותרת
                            cell.font = font_tahoma_header
                            cell.fill = fill_header
                        else:  # שאר הנתונים
                            cell.font = font_tahoma_regular

                # הרחבת העמודות שייראה אסתטי
                worksheet.column_dimensions['A'].width = 20
                worksheet.column_dimensions['B'].width = 15

            st.success("✨ עיבוד הקובץ הסתיים בהצלחה!")

            # דינמיות לשם הקובץ: חילוץ השם המקורי והוספת הטקסט החדש
            original_name, _ = os.path.splitext(uploaded_file.name)
            new_file_name = f"{original_name}_מוכן_לפורטל.xlsx"

            # כפתור הורדה
            st.download_button(
                label="⬇️ הורד קובץ מוכן לפורטל",
                data=buffer.getvalue(),
                file_name=new_file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"שגיאה בעיבוד הקובץ. ודאו שהקובץ תקין ושיש בו נתונים החל מהשורה השנייה.")

# --- אזור הלוגואים בתחתית הדף ---
st.write("")
st.write("")
st.markdown("---")

col1, col2 = st.columns([3, 1])

with col1:
    if os.path.exists("ATEKA_Logo_He.png"):
        st.image("ATEKA_Logo_He.png", width=160)
    else:
        st.write("לוגו אטקה (חסר קובץ ATEKA_Logo_He.png)")

with col2:
    if os.path.exists("yg_logo.png"):
        st.image("yg_logo.png", width=80)
    else:
        st.write("לוגו YG (חסר קובץ yg_logo.png)")