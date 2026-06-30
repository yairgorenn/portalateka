import pandas as pd
import io
import os
import re
from decimal import Decimal, InvalidOperation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from db_handler import find_sku_in_db

# =========================
# הגדרות כלליות - קל לשינוי
# =========================
SKU_LENGTH = 9

# שמות לשוניות
PRIORITY_SHEET_NAME = 'העתקה לפריוריטי'
PORTAL_SHEET_NAME = 'טעינה לפורטל'

# עיצוב כללי
EXCEL_FONT_NAME = 'Tahoma'
EXCEL_FONT_SIZE = 14
HEADER_FONT_SIZE = 14
BODY_ROW_HEIGHT = 24
HEADER_ROW_HEIGHT = 28
ALIGNMENT_DATA = Alignment(horizontal='center', vertical='center')
ALIGNMENT_HEADER = Alignment(horizontal='center', vertical='center')

# צבעים
WARNING_FILL_COLOR = 'FFFF99'   # צהוב בהיר לשורות בעייתיות
SUCCESS_FILL_COLOR = 'E6FFCC'   # ירוק בהיר לשדות להעתקה
HEADER_FILL_COLOR = 'D9EAF7'    # כחול בהיר לכותרות

# פורמט כמויות
# הכמויות נשמרות כטקסט כדי למנוע מאקסל להציג 9 בתור 9.0 או לשנות פורמט בהדבקה
EXCEL_TEXT_FORMAT = '@'

# רוחב עמודות - לשונית העתקה לפריוריטי
PRIORITY_COL_WIDTHS = {
    'A': 30,  # מק"ט
    'B': 20,   # רווח
    'C': 16,  # כמות
    'D': 55,  # הערות מערכת
}

# רוחב עמודות - לשונית טעינה לפורטל
PORTAL_COL_WIDTHS = {
    'A': 30,  # מק"ט
    'B': 16,  # כמות
    'C': 55,  # הערות מערכת
}


def _excel_font(bold=False):
    """יוצר פונט אחיד לכל הקובץ."""
    return Font(name=EXCEL_FONT_NAME, size=EXCEL_FONT_SIZE, bold=bold)


def _apply_column_widths(ws, widths):
    """הגדרת רוחב עמודות לפי מילון."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _style_header_row(ws, max_col, border, header_fill):
    """עיצוב שורת כותרות."""
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

    for cell in ws[1][:max_col]:
        cell.font = _excel_font(bold=True)
        cell.fill = header_fill
        cell.alignment = ALIGNMENT_HEADER
        cell.border = border


def _style_body_rows(ws, min_col, max_col, border):
    """עיצוב בסיסי לכל שורות הנתונים."""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=min_col, max_col=max_col):
        ws.row_dimensions[row[0].row].height = BODY_ROW_HEIGHT
        for cell in row:
            cell.font = _excel_font()
            cell.alignment = ALIGNMENT_DATA
            cell.border = border


def _normalize_quantity(raw_qty):
    """
    מנרמלת כמות לפורמט שלם ונקי להדבקה לפריוריטי.

    חוקים:
    - כמות שלמה נשמרת ללא נקודה עשרונית: 9, 9.0, 9.00 -> '9'
    - כמות עשרונית אמיתית נשארת כפי שהגיעה: 9.3 -> '9.3', אבל מסומנת כלא תקינה
    - ערך ריק / טקסט לא מספרי נשאר כפי שהוא ומסומן כלא תקין
    """
    if raw_qty is None:
        return "", False

    qty_text = str(raw_qty).strip()

    if not qty_text or qty_text.lower() in {"nan", "none", "null"}:
        return "", False

    # אם יש פסיקים של אלפים בלבד, למשל 1,000 או 12,000 - נסיר אותם.
    # אם זה משהו אחר כמו 9,3, נשאיר כמו שהוא ונגדיר כלא תקין.
    qty_for_check = qty_text
    if "," in qty_for_check:
        if re.fullmatch(r"\d{1,3}(,\d{3})+(\.\d+)?", qty_for_check):
            qty_for_check = qty_for_check.replace(",", "")
        else:
            return qty_text, False

    try:
        qty_decimal = Decimal(qty_for_check)
    except InvalidOperation:
        return qty_text, False

    # כמות שלילית לא אמורה להיכנס להזמנת לקוח רגילה
    if qty_decimal < 0:
        return qty_text, False

    # כמות שלמה - מחזירים בלי נקודה עשרונית
    if qty_decimal == qty_decimal.to_integral_value():
        return str(int(qty_decimal)), True

    # כמות עשרונית אמיתית - משאירים את הערך המקורי אבל מסמנים כשגיאה
    return qty_text, False


def _set_text_format(ws, cells):
    """הגדרת תאים כטקסט כדי לשמור על הערך כמו שהוא מוצג."""
    for cell_ref in cells:
        ws[cell_ref].number_format = EXCEL_TEXT_FORMAT


def process_unified_data(items_list, original_file_name):
    """
    מקבלת רשימת פריטים ומייצרת קובץ אקסל בעל שתי לשוניות:
    1. העתקה לפריוריטי: מק"ט | ריק (רווח צר) | כמות | הערות
    2. טעינה לפורטל: מבנה רגיל (מק"ט | כמות | הערות)
    """
    if not items_list:
        return None, None, [], "❌ שגיאה: לא נמצאו פריטים תקינים לפענוח במסמך."

    priority_data = []
    portal_data = []
    warnings = []

    for item in items_list:
        orig_sku = str(item.get('sku', '')).strip()
        orig_qty = str(item.get('qty', '')).strip()

        # סינון שורות ריקות לחלוטין
        if not orig_sku and not orig_qty:
            continue

        clean_sku = orig_sku.upper().replace(" ", "").replace("-", "")
        clean_qty, is_qty_valid = _normalize_quantity(orig_qty)

        final_sku = orig_sku
        status_note = ""
        is_valid = True

        # === 1. חיפוש חכם במסד הנתונים ===
        found_sku = find_sku_in_db(clean_sku) if clean_sku else None

        if found_sku:
            final_sku = str(found_sku).zfill(SKU_LENGTH)
        else:
            status_note = "❌ מק\"ט לא מוכר "
            is_valid = False
            warnings.append(item)

        # === 2. בדיקת כמות ===
        if not is_qty_valid:
            status_note += " | ⚠️ כמות לא תקינה"
            is_valid = False
            if item not in warnings:
                warnings.append(item)

        # === 3. הרכבת הנתונים לגיליון פריוריטי ===
        priority_data.append({
            'מק"ט': final_sku,
            'תיאור (רווח)': "",  # עמודה B נשארת ריקה
            'כמות': clean_qty,
            'הערות מערכת': status_note,
            'is_valid': is_valid  # שדה עזר לעיצוב הצבעים - מוסר לפני הכתיבה לאקסל
        })

        # === 4. הרכבת הנתונים לגיליון הפורטל ===
        portal_data.append({
            'מק"ט': final_sku,
            'כמות': clean_qty,
            'הערות מערכת': status_note
        })

    df_priority = pd.DataFrame(priority_data)
    df_portal = pd.DataFrame(portal_data)

    buffer = io.BytesIO()

    # יצירת קובץ אקסל רב-לשוניות
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # מסירים את עמודת ה-is_valid מגיליון הפריוריטי לפני הכתיבה לקובץ
        df_priority.drop(columns=['is_valid']).to_excel(writer, sheet_name=PRIORITY_SHEET_NAME, index=False)
        df_portal.to_excel(writer, sheet_name=PORTAL_SHEET_NAME, index=False)

        workbook = writer.book

        fill_warning = PatternFill(start_color=WARNING_FILL_COLOR, end_color=WARNING_FILL_COLOR, fill_type='solid')
        fill_success = PatternFill(start_color=SUCCESS_FILL_COLOR, end_color=SUCCESS_FILL_COLOR, fill_type='solid')
        fill_header = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type='solid')

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # === עיצוב גיליון 1: העתקה לפריוריטי ===
        ws_priority = workbook[PRIORITY_SHEET_NAME]
        ws_priority.sheet_view.rightToLeft = True
        ws_priority.freeze_panes = 'A2'
        ws_priority.auto_filter.ref = ws_priority.dimensions

        _style_header_row(ws_priority, max_col=4, border=thin_border, header_fill=fill_header)
        _style_body_rows(ws_priority, min_col=1, max_col=4, border=thin_border)

        # צביעה לפי תקינות ושמירת פורמט טקסט למק"ט וכמות
        for row_idx, row in enumerate(ws_priority.iter_rows(min_row=2, max_row=ws_priority.max_row, min_col=1, max_col=4)):
            current_is_valid = priority_data[row_idx]['is_valid']

            # A = מק"ט, C = כמות. נשמרים כטקסט כדי למנוע שינויי פורמט של Excel.
            row[0].number_format = EXCEL_TEXT_FORMAT
            row[2].number_format = EXCEL_TEXT_FORMAT

            if not current_is_valid:
                for cell in row:
                    cell.fill = fill_warning
            else:
                # אם השורה תקינה, נצבע רק את עמודות A, B, C בירוק - זה האזור שמעתיקים לפריוריטי
                row[0].fill = fill_success
                row[1].fill = fill_success
                row[2].fill = fill_success

        _apply_column_widths(ws_priority, PRIORITY_COL_WIDTHS)

        # === עיצוב גיליון 2: טעינה לפורטל ===
        ws_portal = workbook[PORTAL_SHEET_NAME]
        ws_portal.sheet_view.rightToLeft = True
        ws_portal.freeze_panes = 'A2'
        ws_portal.auto_filter.ref = ws_portal.dimensions

        _style_header_row(ws_portal, max_col=3, border=thin_border, header_fill=fill_header)
        _style_body_rows(ws_portal, min_col=1, max_col=3, border=thin_border)

        for row in ws_portal.iter_rows(min_row=2, max_row=ws_portal.max_row, min_col=1, max_col=3):
            # A = מק"ט, B = כמות. נשמרים כטקסט כדי למנוע שינויי פורמט של Excel.
            row[0].number_format = EXCEL_TEXT_FORMAT
            row[1].number_format = EXCEL_TEXT_FORMAT

            note_cell = row[2].value
            if note_cell and ("❌" in str(note_cell) or "⚠️" in str(note_cell)):
                for cell in row:
                    cell.fill = fill_warning

        _apply_column_widths(ws_portal, PORTAL_COL_WIDTHS)

    buffer.seek(0)
    original_name, _ = os.path.splitext(original_file_name)
    new_file_name = f"{original_name}_מוכן_לפורטל.xlsx"

    return buffer, new_file_name, warnings, None


def process_excel(uploaded_file, original_file_name):
    """
    פונקציית מעטפת לאקסל: קוראת את הקובץ וממירה אותו לרשימת מילונים
    שמתאימה לפונקציית הליבה.
    """
    try:
        if original_file_name.endswith('.csv'):
            df = pd.read_csv(uploaded_file, skiprows=1, header=None)
        else:
            df = pd.read_excel(uploaded_file, skiprows=1, header=None)

        if df.shape[1] < 2:
            return None, None, [], "❌ שגיאה: הקובץ חייב להכיל לפחות שתי עמודות (מק\"ט וכמות)."

        items_list = []
        for index, row in df.iterrows():
            items_list.append({
                'row_num': index + 1,
                'sku': str(row[0]),
                'qty': str(row[1]) if pd.notna(row[1]) else ""
            })

        return process_unified_data(items_list, original_file_name)

    except Exception as e:
        return None, None, [], f"❌ שגיאה בלתי צפויה: {e}"
